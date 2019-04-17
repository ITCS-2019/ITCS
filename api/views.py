from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import JsonResponse

from rest_framework import generics

from mariner.models import Certificate, TrainigOrganisation, TrainigDirections, Sailor
from .serializers import TrainigDirectionsSerializer

from django.http import HttpResponse

# from django.core.files.storage import FileSystemStorage
from django.template.loader import render_to_string
from weasyprint import HTML

import xlwt

import openpyxl

import datetime


class ListTrainigDirectionsView(generics.ListAPIView):
	queryset = TrainigDirections.objects.all()
	serializer_class = TrainigDirectionsSerializer

@login_required(login_url="login/")
def certificates(request):
	if request.user.groups.all()[0].name == 'НТЗ':
		trainigOrganisation = TrainigOrganisation.objects.get(organisation_name=request.user.profile.organization_name)
		certs = list(Certificate.objects.filter(trainigOrganisation=trainigOrganisation).values())
		data =  dict()
		data['certificates'] = certs
		return JsonResponse(data)
	elif request.user.groups.all()[0].name == 'Інспектор':
		certs = list(Certificate.objects.exclude(status=0).values())
		data =  dict()
		data['certificates'] = certs
		return JsonResponse(data)
	else:
		certs = list(Certificate.objects.all().values())
		data =  dict()
		data['certificates'] = certs
		return JsonResponse(data)

@login_required(login_url="login/")
def issuedCerts(request):
	if request.user.groups.all()[0].name == 'НТЗ':
		trainigOrganisation = TrainigOrganisation.objects.get(organisation_name=request.user.profile.organization_name)
		certifications = list(Certificate.objects.filter(trainigOrganisation=trainigOrganisation).filter(status=2).values())
		data =  dict()
		data['certificates'] = certifications
		return JsonResponse(data)
	else:
		certs = list(Certificate.objects.filter(status=2).values())
		data =  dict()
		data['certificates'] = certs
		return JsonResponse(data)

@login_required(login_url="login/")
def changeTrinigDirectionStatus(request):


def changeCertNumber(request):
	certID = request.GET.get('certID')
	certNumber = request.GET.get('certNumber')
	hasError = False
	errorMessage = "No Error"
	if certNumber != None:
		if certNumber != "":
			certificate = Certificate.objects.get(id=certID)
			certificate.certf_number = certNumber
			certificate.save()
		else:
			hasError = True
			errorMessage = "Number is Empty"
	else:
			hasError = True
			errorMessage = "Number is None"
	data = {
		'error' : hasError,
		'error_message' : errorMessage,
	}
	return JsonResponse(data)


@login_required(login_url="login/")
def changeToReviewStatus(request):
	certIDsList = request.GET.get('certIDs').split(',')
	hasError = False
	errorMessage = "No Error"
	if certIDsList != '':
		certsInChange = Certificate.objects.filter(pk__in=certIDsList)
		for cert in certsInChange:
			if cert.status == 0:
				cert.status = 1
				cert.save()
		data = {
			'error' : hasError,
			'error_message' : errorMessage,
		}
		return JsonResponse(data)
	else:
		hasError = True
		errorMessage = "Empty certIDs"
		data = {
			'error' : hasError,
			'error_message' : errorMessage,
		}
		return JsonResponse(data)

@login_required(login_url="login/")
def removeDraftCerts(request):
	certIDsList = request.GET.get('certIDs').split(',')
	hasError = False
	errorMessage = "No Error"
	if certIDsList != '':
		certsInChange = Certificate.objects.filter(pk__in=certIDsList)
		for cert in certsInChange:
			if cert.status == 0:
				cert.delete()
		data = {
			'error' : hasError,
			'error_message' : errorMessage,
		}
		return JsonResponse(data)
	else:
		hasError = True
		errorMessage = "Empty certIDs"
		data = {
			'error' : hasError,
			'error_message' : errorMessage,
		}
		return JsonResponse(data)

@login_required(login_url="login/")
def exportXLS(request):
	print(request.GET.get('exportType'))
	certIDsList = request.GET.get('certIDs').split(',')
	# certsQuerySet = Certificate.objects.filter(pk__in=certIDsList)
	if request.GET.get('exportType') == 'Register':
		rows = Certificate.objects.filter(pk__in=certIDsList).values_list(
		'certf_number', 'first_name_en', 'last_name_en', 'last_name_ukr', 'first_name_ukr', 'second_name_ukr',
		'born', 'date_of_issue', 'valid_date')
		response = HttpResponse(content_type='application/ms-excel')
		response['Content-Disposition'] = 'attachment; filename="certificates.xls"'
		wb = xlwt.Workbook(encoding='utf-8')
		ws = wb.add_sheet('Сертифікати')
		
		row_num = 0
		font_style = xlwt.XFStyle()
		font_style.font.bold = True
		columns = ['Номер документу', 'Name', 'SurName', 'Прізвище','Ім\'я', 'По батькові', 'Дата народження', 'Дата видачі', 'Дійсний до',]
		for col_num in range(len(columns)):
			ws.write(row_num, col_num, columns[col_num], font_style)
		# Sheet body, remaining rows
		font_style = xlwt.XFStyle()
		font_style.num_format_str = 'dd.mm.yyyy'
		
		if rows.count() > 0:
			for row in rows:
				row_num += 1
				for col_num in range(len(row)):
					ws.write(row_num, col_num, row[col_num], font_style)
			wb.save(response)
			return response
		else:
			return HttpResponse(status=204)
	elif request.GET.get('exportType') == 'GiveCerts':
		#need filter(status__startswith=1) check???
		certsInReview = Certificate.objects.filter(pk__in=certIDsList).exclude(certf_number__isnull=True)
		if certsInReview.count() > 0:
			now = datetime.datetime.now()
			for certificate in certsInReview:
				if certificate.certf_number != None:
					if certificate.certf_number != '':
						print('////////////////////')
						#TODO: create hash sum
						ntzNumberStr = "{}/{}{}{}{}".format(
							certificate.certf_number, 
							certificate.trainigOrganisation.organisation_id, 
							now.month,
							now.year,
							certificate.training_direction.id)
						print(ntzNumberStr)
						certificate.ntz_number = ntzNumberStr
						certificate.status = 2
						certificate.save()
					else:
						print('Cert number is Empty')
				else:
					print('Cert number is None')
			return redirect('crm_home')
		else:
			#ErrorMessage: "Zero Certs"
			print('Zero Certs')
			return HttpResponse(status=204)
	else:
		#ErrorMessage: "Incorrect Export Type"
		print('Incorrect Export Type')
		return HttpResponse(status=204)

@login_required(login_url="login/")
def exportToPrint(request):
	print(request.GET.get('exportType'))
	certIDsList = request.GET.get('certIDs').split(',')
	if request.GET.get('exportType') == 'XLSExp':
		rows = Certificate.objects.filter(pk__in=certIDsList).values_list(
		'certf_number', 'first_name_en', 'last_name_en', 'last_name_ukr', 'first_name_ukr', 'second_name_ukr',
		'born', 'date_of_issue', 'valid_date')
		response = HttpResponse(content_type='application/ms-excel')
		response['Content-Disposition'] = 'attachment; filename="certificates.xls"'
		wb = xlwt.Workbook(encoding='utf-8')
		ws = wb.add_sheet('Сертифікати')
		
		row_num = 0
		font_style = xlwt.XFStyle()
		font_style.font.bold = True
		columns = ['Номер документу', 'Name', 'SurName', 'Прізвище','Ім\'я', 'По батькові', 'Дата народження', 'Дата видачі', 'Дійсний до',]
		for col_num in range(len(columns)):
			ws.write(row_num, col_num, columns[col_num], font_style)
		# Sheet body, remaining rows
		font_style = xlwt.XFStyle()
		font_style.num_format_str = 'dd.mm.yyyy'
		
		if rows.count() > 0:
			for row in rows:
				row_num += 1
				for col_num in range(len(row)):
					ws.write(row_num, col_num, row[col_num], font_style)
			wb.save(response)
			return response
		else:
			return HttpResponse(status=204)
	elif request.GET.get('exportType') == 'PdfExp':
		certifications = Certificate.objects.filter(pk__in=certIDsList)
		html_string = render_to_string('printTable.html', {'certifications': certifications})
		html = HTML(string=html_string).write_pdf()
		response = HttpResponse(html, content_type='application/pdf')
		response['Content-Disposition'] = "inline; filename=certificates.pdf"
		return response
	elif request.GET.get('exportType') == 'PrintExp':
		return HttpResponse(status=204)
	else:
		print('Incorrect Export Type')
		return HttpResponse(status=204)

@login_required(login_url="login/")
def uploadXLS(request):
	if "GET" == request.method:
		print('GET request')
		return render(request, "crm_xlsImport.html", {})
	else:
		print('load file')
		excel_file = request.FILES["excel_file"]
		# TODO: validations check extension or file size
		wb = openpyxl.load_workbook(excel_file, read_only=False, keep_vba=False, data_only=False, keep_links=True)
		# getting a particular sheet by name out of many sheets
		worksheet = wb[0]
		print(worksheet)
		excel_data = list()
		# iterating over the rows and
		# getting value from each cell in row
		for row in worksheet.iter_rows():
			row_data = list()
			for cell in row:
				row_data.append(str(cell.value))
			excel_data.append(row_data)
		print(excel_data)
		return render(request, 'crm_xlsImport.html', {"excel_data":excel_data})
		#return HttpResponse(status=204)