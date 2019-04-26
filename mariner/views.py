from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import JsonResponse

from accounts.models import Profile

from regulations.models import RegulationDoc
from mariner.models import *
from mariner.forms import UserForm, ProfileForm, RegulationForm, CertificationForm, CertificationFormPDF, SailorForm, TrainigOrganisationForm, TrainigDirectionsForm, UploadXLSForm

import datetime

from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML

import xlwt
import re
import xlrd
import openpyxl
import os
#from transliterate import translit
from transliterate.base import TranslitLanguagePack, registry
from transliterate import translit, get_available_language_codes

@login_required(login_url="login/")
def crm_home(request):
	profile, created = Profile.objects.get_or_create(user=request.user)
	if request.user.groups.all()[0].name == 'НТЗ':
		if request.user.profile.organization_name == '':
			return redirect('update_profile')
		else:
			sailorsCount = Sailor.objects.all().count
			trainigOrganisation = TrainigOrganisation.objects.get(organisation_name=request.user.profile.organization_name)
			trainigDirectionsCount = trainigOrganisation.directions.count
			certCount = trainigOrganisation.trained.filter(status__startswith=2).count
			certsInDraftCount = trainigOrganisation.trained.filter(status__startswith=0).count
			certsInReviewCount = trainigOrganisation.get_certInReview().count
			context = {'sailorsCount': sailorsCount, 'certCount': certCount, 'trainigDirectionsCount': trainigDirectionsCount, 'certsInDraftCount': certsInDraftCount, 'certsInReviewCount': certsInReviewCount,}
			return render(request, "crm_dashboard.html", context)
	else:
		sailorsCount = Sailor.objects.all().count
		certCount = Certificate.objects.filter(status__startswith=2).count
		certsInReviewCount = Certificate.objects.filter(status__startswith=1).count
		trainigDirectionsCount = TrainigDirections.objects.all().count
		# certsInReview = Certificate.objects.filter(status__startswith=1)
		trainigOrganisations = TrainigOrganisation.objects.all()
		context = {'trainigOrganisations': trainigOrganisations, 'sailorsCount': sailorsCount, 'certCount': certCount, 'trainigDirectionsCount': trainigDirectionsCount, 'certsInReviewCount': certsInReviewCount,}
		return render(request, "crm_dashboard.html", context)

@login_required(login_url="login/")
def crm_profile(request):
	profile, created = Profile.objects.get_or_create(user=request.user)
	return render(request, "crm_profile.html",)

@login_required(login_url="login/")
def update_profile(request):
    if request.method == 'POST':
        user_form = UserForm(request.POST, instance=request.user)
        profile_form = ProfileForm(request.POST, instance=request.user.profile)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            return redirect('crm_profile')
    else:
        user_form = UserForm(instance=request.user)
        profile_form = ProfileForm(instance=request.user.profile)
    return render(request, 'crm_editProfile.html', {
        'user_form': user_form,
        'profile_form': profile_form
    })

#/////////////////TrainigOrganisation//////////////////////
@login_required(login_url="login/")
def crm_trainigOrganisations(request):
	trainigOrganisations = TrainigOrganisation.objects.all()
	context = {'trainigOrganisations': trainigOrganisations,}
	return render(request, "crm_trainigOrganisations.html", context)

@login_required(login_url="login/")
def crm_trainigOrganisationView(request, name):
	trainigOrganisation = TrainigOrganisation.objects.get(organisation_name=name)
	context = {'trainigOrganisation': trainigOrganisation,}
	return render(request, "crm_trainigOrganisationDetail.html", context)

@login_required(login_url="login/")
def crm_trainigOrganisationDirectionView(request, organisation_name, direction_name):
	trainigOrganisation = TrainigOrganisation.objects.get(organisation_name=name)
	filtredCerts = trainigOrganisation.trained.filter(directions = direction_name)
	print(filtredCerts)
	context = {'trainigOrganisation': trainigOrganisation, 'directionName': direction_name, 'filtredCerts': filtredCerts}
	return render(request, "crm_trainigOrganisationDirectionDetail.html", context)

@login_required(login_url="login/")
def add_trainigOrganisation(request):
	if request.method == "POST":
		form = TrainigOrganisationForm(request.POST)
		if form.is_valid():
			form.save()
			return redirect('crm_trainigOrganisations')
	else:
		form = TrainigOrganisationForm()
	context = {'form': form,}
	return render(request, 'crm_editTrainigOrganisations.html', context)

@login_required(login_url="login/")
def crm_editTrainigOrganisation(request, name, template_name='crm_editTrainigOrganisations.html'):
    trainigOrganisation = TrainigOrganisation.objects.get(organisation_name=name)
    form = TrainigOrganisationForm(request.POST or None, instance=trainigOrganisation)
    if form.is_valid():
        form.save()
        return redirect('crm_trainigOrganisations')
    context = {'form': form,}
    return render(request, template_name, context)
#//////////////////////////////////////////////////////////
@login_required(login_url="login/")
def crm_sailors(request):
	sailors = Sailor.objects.all()
	context = {'sailors': sailors,}
	return render(request, "crm_sailors.html", context)

@login_required(login_url="login/")
def crm_sailorView(request, id):
	sailor = Sailor.objects.get(id=id)
	context = {'sailor': sailor,}
	return render(request, "crm_sailorDetail.html", context)

@login_required(login_url="login/")
def add_sailor(request):
	if request.method == "POST":
		form = SailorForm(request.POST)
		if form.is_valid():
			sailor = form.save(commit=False)
			if sailor.inn is not None:
				dbSailor = Sailor.objects.filter(inn=sailor.inn).first()
				if dbSailor is not None:
					if sailor.inn != dbSailor.inn:
						sailor.save()
						return redirect('crm_sailors')
					else:
						return redirect('crm_sailors')
				else:
					sailor.save()
					return redirect('crm_sailors')
			else:
				sailor, created = Sailor.objects.get_or_create(
				first_name_en = sailor.first_name_en,
				last_name_en = sailor.last_name_en,
				last_name_ukr = sailor.last_name_ukr,
				first_name_ukr = sailor.first_name_ukr,
				second_name_ukr = sailor.second_name_ukr,
				born = sailor.born,
				)
				if created:
					sailor.save()
				return redirect('crm_sailors')
	else:
		form = SailorForm()
	context = {'form': form,}
	return render(request, 'crm_editSailor.html', context)

@login_required(login_url="login/")
def crm_editSailor(request, id, template_name='crm_editSailor.html'):
    sailor = Sailor.objects.get(id=id)
    form = SailorForm(request.POST or None, instance=sailor)
    if form.is_valid():
        form.save()
        return redirect('crm_sailors')
    context = {'form': form,}
    return render(request, template_name, context)

#//////////////////////////////////////////////////////////
@login_required(login_url="login/")
def crm_trainigDirections(request):
	if request.user.groups.all()[0].name == 'НТЗ':
		trainigOrganisation = TrainigOrganisation.objects.get(organisation_name=request.user.profile.organization_name)
		trainigDirections = trainigOrganisation.directions.all()
		context = {'trainigDirections':trainigDirections}
		return render(request, "crm_trainigDirections.html", context)
	else:
		trainigDirections = TrainigDirections.objects.all()
		context = {'trainigDirections':trainigDirections}
		return render(request, "crm_trainigDirections.html", context)

@login_required(login_url="login/")
def new_trainigDirection(request):
	if request.method == "POST":
		form = TrainigDirectionsForm(request.POST)
		if form.is_valid():
			trainigDirection = form.save(commit=False)
			trainigDirection, created = TrainigDirections.objects.get_or_create(
				direction_title = trainigDirection.direction_title,
				level = trainigDirection.level,
				allow_functions = trainigDirection.allow_functions,
				price_id = trainigDirection.price_id,
				price = trainigDirection.price,
			)
			# if created:#a new direction
			# 	trainigDirection.save()
			return redirect('crm_trainigDirections')
	else:
		form = TrainigDirectionsForm()
	context = {'form': form,}
	return render(request, 'crm_editTrainigDirections.html', context)

@login_required(login_url="login/")
def crm_editTrainigDirection(request, direction_id, template_name='crm_editTrainigDirections.html'):
    direction = TrainigDirections.objects.get(id=direction_id)
    form = TrainigDirectionsForm(request.POST or None, instance=direction)
    if form.is_valid():
        form.save()
        return redirect('crm_trainigDirections')
    context = {'form': form,}
    return render(request, template_name, context)

#//////////////////////////////////////////////////////////
@login_required(login_url="login/")
def crm_regulations(request):
	regulations = RegulationDoc.objects.all()
	context = {'regulations': regulations,}
	return render(request, "crm_regulations.html", context)

@login_required(login_url="login/")
def new_regulation(request):
	if request.method == "POST":
		form = RegulationForm(request.POST)
		if form.is_valid():
			regulation = form.save(commit=False)
			regulation.user = request.user
			regulation.save()
			return redirect('crm_regulations')
	else:
		form = RegulationForm()
	context = {'form': form,}
	return render(request, 'crm_editRegulation.html', context)

@login_required(login_url="login/")
def crm_editRegulation(request, number, template_name='crm_editRegulation.html'):
    regulation = RegulationDoc.objects.get(number=number)
    form = RegulationForm(request.POST or None, instance=regulation)
    if form.is_valid():
        form.save()
        return redirect('crm_regulations')
    context = {'form': form,}
    return render(request, template_name, context)

#//////////////////////////////////////////////////////////
@login_required(login_url="login/")
def crm_certification(request):
	if request.user.groups.all()[0].name == 'НТЗ':
		trainigOrganisation = TrainigOrganisation.objects.get(organisation_name=request.user.profile.organization_name)
		certifications = Certificate.objects.filter(trainigOrganisation=trainigOrganisation)
		context = {'certifications': certifications,}
		return render(request, "crm_certification.html", context)
	elif request.user.groups.all()[0].name == 'Інспектор':
		certs = Certificate.objects.exclude(status=0)
		context = {'certifications': certs,}
		return render(request, "crm_certification.html", context)
	else:
		certs = Certificate.objects.all()
		context = {'certifications': certs,}
		return render(request, "crm_certification.html", context)

@login_required(login_url="login/")
def new_certification(request):
	if request.method == "POST":
		form = CertificationForm(request.user, 0, request.POST)
		if form.is_valid():
			certification = form.save(commit=False)
			if certification.inn is not None:
				dbSailor = Sailor.objects.filter(inn=certification.inn).first()
				if dbSailor is not None:
					if certification.inn != dbSailor.inn:
						sailor = Sailor()
						sailor.first_name_en = certification.first_name_en
						sailor.last_name_en = certification.last_name_en
						sailor.last_name_ukr = certification.last_name_ukr
						sailor.first_name_ukr = certification.first_name_ukr
						sailor.second_name_ukr = certification.second_name_ukr
						sailor.born = certification.born
						sailor.died = certification.died
						sailor.inn = certification.inn
						sailor.save()
					else:
						sailor = Sailor()
						sailor = dbSailor
						sailor.save()
				else:
					sailor = Sailor()
					sailor.first_name_en = certification.first_name_en
					sailor.last_name_en = certification.last_name_en
					sailor.last_name_ukr = certification.last_name_ukr
					sailor.first_name_ukr = certification.first_name_ukr
					sailor.second_name_ukr = certification.second_name_ukr
					sailor.born = certification.born
					sailor.inn = certification.inn
					sailor.save()
			else:
				sailor, created = Sailor.objects.get_or_create(
				first_name_en = certification.first_name_en,
				last_name_en = certification.last_name_en,
				last_name_ukr = certification.last_name_ukr,
				first_name_ukr = certification.first_name_ukr,
				second_name_ukr = certification.second_name_ukr,
				born = certification.born,
				)
				if created:
					sailor.save()
			certification.sailor = sailor
			if request.user.groups.all()[0].name == 'НТЗ':
				organisation = TrainigOrganisation.objects.get(organisation_name=request.user.profile.organization_name)
				certification.trainigOrganisation = organisation
			#certification, created = Certificate.objects.get_or_create(
				#sailor = certification.sailor,
				#trainigOrganisation = certification.trainigOrganisation,
				#date_of_issue = certification.date_of_issue,
				#valid_date = certification.valid_date,
				#valid_type = certification.valid_type,
				#training_direction = certification.training_direction,
				#).first()
			#if created:
			certification.save()
			return redirect('crm_certification')
	else:
		certifStatus = 0
		form = CertificationForm(request.user, certifStatus)
	sailors = Sailor.objects.all()
	context = {'form': form, 'sailors': sailors, 'statusNum': 0,}
	return render(request, 'crm_editCertificate.html', context)

@login_required(login_url="login/")
def crm_editCertification(request, id, template_name='crm_editCertificate.html'):
    certificate = Certificate.objects.get(id=id)
    form = CertificationForm(request.user, certificate.status, request.POST or None, instance=certificate)
    if form.is_valid():
        form.save()
        if request.user.groups.all()[0].name == 'НТЗ':
        	organisation = TrainigOrganisation.objects.get(organisation_name=request.user.profile.organization_name)
        	certificate.trainigOrganisation = organisation
        	certificate.save()
        	return redirect('crm_certification')
        else:
        	return redirect('crm_home')
    sailors = Sailor.objects.all()
    context = {'form': form, 'sailors': sailors, 'statusNum': certificate.status,}
    return render(request, template_name, context)

#//////////////////////////////////////////////////////////
@login_required(login_url="login/")
def crm_exportCertifsForm(request):
	if request.user.groups.all()[0].name == 'НТЗ':
		trainigOrganisation = TrainigOrganisation.objects.get(organisation_name=request.user.profile.organization_name)
		certifications = Certificate.objects.filter(trainigOrganisation=trainigOrganisation).filter(status=2)
		context = {'certifications': certifications,}
		return render(request, "crm_certification.html", context)
	else:
		certs = Certificate.objects.filter(status=2)
		context = {'certifications': certs,}
		return render(request, "crm_exportCertfsForm.html", context)

@login_required(login_url="login/")
def crm_certForm1(request):
	if request.method == "POST":
		form = CertificationFormPDF(request.POST)
		if form.is_valid():
			cd = form.cleaned_data
			startDate = cd.get('startDate')
			endDate = cd.get('endDate')
			certifications = Certificate.objects.all()
			html_string = render_to_string('request_form1.html', {'certifications': certifications})
			html = HTML(string=html_string).write_pdf()
			response = HttpResponse(html, content_type='application/pdf')
			response['Content-Disposition'] = "inline; filename=certificates.pdf"
			return response
			#return redirect('crm_certification')
	else:
		form = CertificationFormPDF()
	return render(request, 'crm_form1.html', {'form': form,})
#//////////////////////////////////////////////////////////
def load_certifs(request):
	dateFrom = request.GET.get('dateFrom')
	dateTo = request.GET.get('dateTo')
	directionName = request.GET.get('directionName')
	
	certif =  list(Certificate.objects.filter(status=2).values())
	
	data =  dict()
	data['certificates'] = certif
	return JsonResponse(data)

def load_directions(request):
	organisationName = request.GET.get('organisation')
	trainigOrganisation = TrainigOrganisation.objects.get(organisation_name=organisationName)
	directionsOfOrganization = trainigOrganisation.directions.all()
	
	data = {
		'directions': list(directionsOfOrganization.values('direction_title','level'))
	}
	print(data)
	return JsonResponse(data)

class UkrTranslitLanguagePack(TranslitLanguagePack):
	language_code = "ukrTranslit"
	language_name = "UkrTranslit"
	mapping = (
		u"абвгґдезиіїйклмнопрстуфАБВГҐДЕЗИІЙКЛМНОПРСТУФ",
		u"abvhgdezyiiiklmnoprstufABVHGDEZYIYKLMNOPRSTUF",
	)
	pre_processor_mapping = {
	u"Є": u"Ye",
	u"є": u"ie",
	u"Ж": u"Zh",
	u"ж": u"zh",
	u"Ї": u"Yi",
	u"Х": u"Kh",
	u"х": u"kh",
	u"Ц": u"Ts",
	u"ц": u"ts",
	u"Ч": u"Ch",
	u"ч": u"ch",
	u"Ш": u"Sh",
	u"ш": u"sh",
	u"Щ": u"Shch",
	u"щ": u"shch",
	u"Ю": u"Yu",
	u"ю": u"iu",
	u"Я": u"Ya",
	u"я": u"ia",
	u"ь": u"",
	u"ъ": u"",
	u"\'": u"",
	}

registry.register(UkrTranslitLanguagePack)

def transliterateField(request):
	
	fieldData = request.GET.get('fieldData')
	# print(translit(fieldData, 'en'))
	print(get_available_language_codes())
	print(translit(fieldData, 'ukrTranslit'))
	fieldData = fieldData.lower().title()
	data = {
	'transliteration': translit(fieldData, 'ukrTranslit'),
	}
	return JsonResponse(data)

def certf_fieldValid(request):
	fieldName = request.GET.get('fieldName')
	fieldData = request.GET.get('fieldData')
	hasError = False
	errorMessage = "No Error"
	# print(fieldName)
	# print(fieldData)
	if fieldData != None:
		if fieldName == "first_name_en":
			if not bool(re.fullmatch('[a-zA-Z\-\'\s]{2,32}( [a-zA-Z\-\'\s]{2,32})?', fieldData)):
				hasError = True
				errorMessage = "Помилка. Поле First Name повинно бути на латиниці. Якщо це кілька слів вони повинні бути розділені пропуском або дефісом. Наприклад: \"Tan Zung\""
		if fieldName == "last_name_en":
			if not bool(re.fullmatch('[a-zA-Z\-\'\s]{2,32}( [a-zA-Z\-\'\s]{2,32})?', fieldData)):
				hasError = True
				errorMessage = "Помилка. Поле Surname повинно бути на латиниці. Якщо це кілька слів вони повинні бути розділені пропуском або дефісом. Наприклад: \"Saltykov-Shchedrin\""
	data = {
		'error' : hasError,
		'error_message' : errorMessage,
	}
	return JsonResponse(data)

#//////////////////////////////////////////////////////////
@login_required(login_url="login/")
def exportToRegister(request, name):
	trainigOrganisation = TrainigOrganisation.objects.get(organisation_name=name)
	rows = trainigOrganisation.get_certInReview().exclude(certf_number__isnull=True).values_list(
		'certf_number', 'first_name_en', 'last_name_en', 'last_name_ukr', 'first_name_ukr', 'second_name_ukr',
		'born', 'date_of_issue', 'valid_date')
	#certsInReview = Certificate.objects.filter(status__startswith=1)
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="certificates.xls"'
	wb = xlwt.Workbook(encoding='utf-8')
	ws = wb.add_sheet('Сертифікати')
	# Sheet header, first row
	row_num = 0
	font_style = xlwt.XFStyle()
	font_style.font.bold = True
	columns = ['Номер документу', 'Name', 'SurName', 'Прізвище','Ім\'я', 'По батькові', 'Дата народження', 'Дата видачі', 'Дійсний до',]
	for col_num in range(len(columns)):
		ws.write(row_num, col_num, columns[col_num], font_style)
	# Sheet body, remaining rows
	font_style = xlwt.XFStyle()
	#font_style.num_format_str = 'dd/mm/yyyy'
	font_style.num_format_str = 'dd.mm.yyyy'
	# rows = certsInReview.objects.all().values_list(
	# 	'first_name_en', 'last_name_en', 'last_name_ukr', 'first_name_ukr', 'second_name_ukr',
	# 	'born', 'date_of_issue', 'valid_date')
	print(rows.count())
	if rows.count() > 0:
		for row in rows:
			row_num += 1
			for col_num in range(len(row)):
				ws.write(row_num, col_num, row[col_num], font_style)
		wb.save(response)
		return response
	else:
		return HttpResponse(status=204)

@login_required(login_url="login/")
def giveCertificates(request, name):
	trainigOrganisation = TrainigOrganisation.objects.get(organisation_name=name)
	certsInReview = trainigOrganisation.get_certInReview().exclude(certf_number__isnull=True)
	#certsInReview.exclude(certf_number__isnull=True).exclude(certf_number__exact='')
	if certsInReview.count() > 0:
		now = datetime.datetime.now()
		for certificate in certsInReview:
			if certificate.certf_number != None:
				if certificate.certf_number != '':
					print('////////////////////')
					#TODO: create hash sum
					ntzNumberStr = "{}/{}{}{}{}".format(
						certificate.certf_number, 
						certificate.trainigOrganisation.organisation_id, 
						now.month,
						now.year,
						certificate.training_direction.id)
					print(ntzNumberStr)
					certificate.ntz_number = ntzNumberStr
					certificate.status = 2
					certificate.save()
		return redirect('crm_home')
	else:
		return HttpResponse(status=204)

# @login_required(login_url="login/")
# def giveOrganizationCertificates(request, name):
# 	trainigOrganisation = TrainigOrganisation.objects.get(organisation_name=name)
# 	organizationCertificates = trainigOrganisation.trained.all()
# 	organizationCertsInReview = organizationCertificates.objects.filter(status__startswith=1)
#//////////////////////////////////////////////////////////
@login_required(login_url="login/")
def generate_pdf(request):
	certifications = Certificate.objects.all()
	html_string = render_to_string('request_form1.html', {'certifications': certifications})
	html = HTML(string=html_string).write_pdf()

	response = HttpResponse(html, content_type='application/pdf')
	response['Content-Disposition'] = "inline; filename=certificates.pdf"
	# result = html.write_pdf(target='/tmp/mypdf.pdf')
	# fs = FileSystemStorage('/tmp')
	# with fs.open('mypdf.pdf') as pdf:
	# 	response = HttpResponse(pdf, content_type='application/pdf')
	# 	response['Content-Disposition'] = 'attachment; filename="mypdf.pdf"'
	# 	return response

	return response

@login_required(login_url="login/")
def uploadXLS(request):
	if "GET" == request.method:
		form = UploadXLSForm(request.user)
		return render(request, 'crm_xlsImport.html', {'form': form})
	else:
		directionID = request.POST.get('directions')
		# if directionID != '0': # if direction not in table
		# 	directionObj = TrainigDirections.objects.get(id=directionID)
		excel_file = request.FILES['file']
		# TODO: validations check extension or file size
		extesion = os.path.splitext(str(request.FILES['file']))[1]
		if extesion == '.xlsx':
			wb = openpyxl.load_workbook(excel_file, read_only=False, keep_vba=False, data_only=False, keep_links=True)
			# getting a particular sheet by name out of many sheets
			worksheet = wb.active
			#print(worksheet)
			excel_data = list()
			# iterating over the rows and
			# getting value from each cell in row
			for row in worksheet.iter_rows():
				row_data = list()
				for cell in row:
					row_data.append(str(cell.value))
				excel_data.append(row_data)
			#print(excel_data)
			return render(request, 'crm_xlsImport.html', {"excel_data":excel_data})
		elif extesion == '.xls':
			rb = xlrd.open_workbook(file_contents=excel_file.read())
			sheet = rb.sheet_by_index(0)
			
			if sheet.ncols == 9: #Table without direction
				if directionID != '0':
					excel_data = list()
					for rownum in range(sheet.nrows):
						row_data = list()
						curCellIndx = 0
						docNumber = ''
						firstNameEn = ''
						lastNameEn = ''
						lastNameUkr = ''
						firstNameUkr = ''
						secondNameUkr = ''
						birthday = ''
						startDate = ''
						endDate = ''
						row = sheet.row_values(rownum)
						for cell in row:
							if rownum != 0:
								if curCellIndx == 0:
									docNumber = cell
								elif curCellIndx == 1:
									firstNameEn = cell
								elif curCellIndx == 2:
									lastNameEn = cell
								elif curCellIndx == 3:
									lastNameUkr = cell
								elif curCellIndx == 4:
									firstNameUkr = cell
								elif curCellIndx == 5:
									secondNameUkr = cell
								elif curCellIndx == 6:
									cellAsDatetime = datetime.datetime(*xlrd.xldate_as_tuple(float(cell), rb.datemode))
									birthday = cellAsDatetime.date().isoformat()
								elif curCellIndx == 7:
									cellAsDatetime = datetime.datetime(*xlrd.xldate_as_tuple(float(cell), rb.datemode))
									startDate = cellAsDatetime.date().isoformat()
								elif curCellIndx == 8:
									print(cell)
									if cell != '':
										cellAsDatetime = datetime.datetime(*xlrd.xldate_as_tuple(float(cell), rb.datemode))
										endDate = cellAsDatetime.date().isoformat()
									else:
										endDate = None
							row_data.append(cell)
							curCellIndx = curCellIndx + 1
						excel_data.append(row_data)
						if rownum != 0:
							sailor, created = Sailor.objects.get_or_create(
								first_name_en = firstNameEn,
								last_name_en = lastNameEn,
								last_name_ukr = lastNameUkr,
								first_name_ukr = firstNameUkr,
								second_name_ukr = secondNameUkr,
								born = birthday,
							)
							if created:
								sailor.save()

							certificate, created = Certificate.objects.get_or_create(
								form_number = docNumber,
								trainigOrganisation = TrainigOrganisation.objects.get(organisation_name=request.user.profile.organization_name),
								training_direction = TrainigDirections.objects.get(id=directionID),
								first_name_en = firstNameEn,
								last_name_en = lastNameEn,
								last_name_ukr = lastNameUkr,
								first_name_ukr = firstNameUkr,
								second_name_ukr = secondNameUkr,
								born = birthday,
								sailor = sailor,
								date_of_issue = startDate,
								valid_date = endDate,
								)
							if created:
								certificate.save()
					return render(request, 'crm_xlsImport.html', {"excel_data":excel_data, "error_message": "Сертифікати добавлени"})
				else:
					#return render(request, 'crm_xlsImport.html', {"error_message": "Таблиця без напрямку. Будласка, оберить напрямок підготовки"})
					form = UploadXLSForm(request.user)
					return render(request, 'crm_xlsImport.html', {'form': form, "error_message": "Таблиця без напрямку. Будласка, оберить напрямок підготовки"})
			elif sheet.ncols == 11: #Table with directions
				if directionID == '0':
					excel_data = list()
					for rownum in range(sheet.nrows):
						row_data = list()
						curCellIndx = 0
						docNumber = ''
						firstNameEn = ''
						lastNameEn = ''
						lastNameUkr = ''
						firstNameUkr = ''
						secondNameUkr = ''
						birthday = ''
						innCell = ''
						startDate = ''
						endDate = ''
						directionIdCell = ''
						row = sheet.row_values(rownum)
						for cell in row:
							if rownum != 0:
								if curCellIndx == 0:
									docNumber = cell
								elif curCellIndx == 1:
									firstNameEn = cell
								elif curCellIndx == 2:
									lastNameEn = cell
								elif curCellIndx == 3:
									lastNameUkr = cell
								elif curCellIndx == 4:
									firstNameUkr = cell
								elif curCellIndx == 5:
									secondNameUkr = cell
								elif curCellIndx == 6:
									cellAsDatetime = datetime.datetime(*xlrd.xldate_as_tuple(float(cell), rb.datemode))
									birthday = cellAsDatetime.date().isoformat()
								elif curCellIndx == 7:
									innCell = cell
								elif curCellIndx == 8:
									cellAsDatetime = datetime.datetime(*xlrd.xldate_as_tuple(float(cell), rb.datemode))
									startDate = cellAsDatetime.date().isoformat()
								elif curCellIndx == 9:
									print(cell)
									if cell != '':
										cellAsDatetime = datetime.datetime(*xlrd.xldate_as_tuple(float(cell), rb.datemode))
										endDate = cellAsDatetime.date().isoformat()
									else:
										endDate = None
								elif curCellIndx == 10:
									directionIdCell = cell
							row_data.append(cell)
							curCellIndx = curCellIndx + 1
						excel_data.append(row_data)
						if rownum != 0:
							sailor, created = Sailor.objects.get_or_create(
								first_name_en = firstNameEn,
								last_name_en = lastNameEn,
								last_name_ukr = lastNameUkr,
								first_name_ukr = firstNameUkr,
								second_name_ukr = secondNameUkr,
								born = birthday,
								inn = innCell
							)
							if created:
								sailor.save()

							certificate, created = Certificate.objects.get_or_create(
								form_number = docNumber,
								trainigOrganisation = TrainigOrganisation.objects.get(organisation_name=request.user.profile.organization_name),
								training_direction = TrainigDirections.objects.get(id=directionID),
								first_name_en = firstNameEn,
								last_name_en = lastNameEn,
								last_name_ukr = lastNameUkr,
								first_name_ukr = firstNameUkr,
								second_name_ukr = secondNameUkr,
								born = birthday,
								inn = innCell,
								sailor = sailor,
								date_of_issue = startDate,
								valid_date = endDate,
								)
							if created:
								certificate.save()
					return render(request, 'crm_xlsImport.html', {"excel_data":excel_data, "error_message": "Сертифікати добавлени"})
				else:
					return render(request, 'crm_xlsImport.html', {"error_message": "Таблиця з напрямками. Будласка, оберіть авто визначення напрямку підготовки"})
			else:
				return render(request, 'crm_xlsImport.html', {"error_message": "Будласка, перевірте таблицю"})
		else:#incorrect file type
			return render(request, 'crm_xlsImport.html', {"error_message": "Будласка, перевірте тип файлу"})